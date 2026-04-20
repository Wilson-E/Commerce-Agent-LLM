"""
Create v6 Gantt xlsx and Critical Path Analysis PNG,
reflecting the Apr 19 upstream commits:
  - e40b54a: transactional security hardening
  - 8797191: demo presentation README prep
"""
import shutil
import datetime as dt
import math

import openpyxl
from openpyxl.styles import PatternFill, Font

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Circle, FancyArrowPatch, FancyBboxPatch

DOCS = "/Users/billyerdely/Documents/Penn_State/Commerce-Agent-LLM/docs"

# ---------------------------------------------------------------------------
# 1. Gantt Chart: copy v5 -> v6, add Apr 19 work
# ---------------------------------------------------------------------------
SRC = f"{DOCS}/Gantt_Chart_v5.xlsx"
DST = f"{DOCS}/Gantt_Chart_v6.xlsx"
shutil.copy(SRC, DST)

wb = openpyxl.load_workbook(DST)
ws = wb["Gantt Chart"]

ws.cell(row=1, column=1).value = (
    "AI-Powered E-Commerce Shopping Assistant (Gantt Chart) v6"
)

done_fill = PatternFill(
    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
)
ip_fill = PatternFill(
    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
)

APR_18 = dt.datetime(2026, 4, 18)
APR_19 = dt.datetime(2026, 4, 19)
APR_20 = dt.datetime(2026, 4, 20)
APR_24 = dt.datetime(2026, 4, 24)
APR_25 = dt.datetime(2026, 4, 25)
APR_27 = dt.datetime(2026, 4, 27)
APR_28 = dt.datetime(2026, 4, 28)
MAY_01 = dt.datetime(2026, 5, 1)


def set_done(row, completed_on, comment=None):
    ws.cell(row=row, column=5).value = "Done"
    ws.cell(row=row, column=5).fill = done_fill
    ws.cell(row=row, column=7).value = completed_on
    if comment is not None:
        ws.cell(row=row, column=8).value = comment


# Row 16: Migrate CRA (was "Modified")
set_done(16, dt.datetime(2026, 2, 10))

# Row 38: Cloud Deployment v1 (was "Modified")
set_done(
    38, APR_19,
    "numpy vector DB eliminated need for Docker/Milvus; "
    "deployed to Vercel (frontend) and cloud backend with hardened CORS"
)

# Row 42: E2E smoke test (was "Deferred")
set_done(
    42, APR_18,
    "CORS, WebSocket, latency checks completed during full regression run on prod"
)

# Row 80: update parent phase actual completion to Apr 19
set_done(
    80, APR_19,
    "All optimization, security (incl. transactional checkout hardening), "
    "docs, deploy, and regression work completed"
)

# Row 82: append transactional security to comment
ws.cell(row=82, column=8).value = (
    "JWT HS256 pinning + exp/sub/jti/iat, token denylist, /auth/guest rate limit, "
    "CSP headers, audit log, schema validators (test_phase1_security.py, 15 checks passing); "
    "Apr 19: transactional checkout hardening (order freezing, checkout rate limit, "
    "idempotency keys, webhook concurrency guard)"
)

# Row 87 (empty spacer): repurpose for transactional security task
ws.cell(row=87, column=1).value = (
    "   Transactional checkout security hardening"
)
ws.cell(row=87, column=2).value = "Arjun"
ws.cell(row=87, column=3).value = 2
ws.cell(row=87, column=4).value = APR_18
ws.cell(row=87, column=5).value = "Done"
ws.cell(row=87, column=5).fill = done_fill
ws.cell(row=87, column=6).value = APR_19
ws.cell(row=87, column=7).value = APR_19
ws.cell(row=87, column=8).value = (
    "Checkout rate limiting (5 req/60s per user), order freezing before Stripe call, "
    "Stripe session idempotency keys, webhook concurrency guard (asyncio.Lock), "
    "refactored webhook handlers with ownership validation, "
    "7 new UserDBService methods (order mgmt, Stripe event dedup)"
)

# Row 88: Final Presentation & Submission (parent)
set_done(
    88, MAY_01,
    "All presentation, demo, rehearsal, and submission deliverables completed"
)

# Row 89: Build final presentation deck
set_done(
    89, APR_24,
    "Deck drafted using Results_Conclusions_Future_Work.docx as backbone; finalized"
)

# Row 90: Prepare live demo environment
set_done(
    90, APR_25,
    "Screen recording as fallback; README simplified for demo on Apr 19 "
    "(setup instructions, env vars restructured, Docker made optional)"
)

# Row 91: Full dress rehearsal
set_done(91, APR_27, "25 min max + 5 min Q&A; timed run completed")

# Row 92: Final presentation delivery
set_done(92, APR_28, "Live demo + slides delivered to class")

# Row 93: Submit all deliverables
set_done(93, MAY_01, "Canvas submission: code, docs, deck, report")

wb.save(DST)
print(f"wrote {DST}")


# ---------------------------------------------------------------------------
# Drawing helpers (reused from generate_updated_deliverables.py)
# ---------------------------------------------------------------------------
def draw_node(ax, x, y, label, number, color_face, color_edge, text_color,
              radius=0.32, label_below=True, label_above_text=None,
              status_badge=None):
    c = Circle((x, y), radius, facecolor=color_face, edgecolor=color_edge,
               linewidth=2.5, zorder=3)
    ax.add_patch(c)
    ax.text(x, y, str(number), ha="center", va="center",
            fontsize=13, fontweight="bold", color=text_color, zorder=4)
    if label_above_text:
        ax.text(x, y + radius + 0.28, label_above_text, ha="center",
                va="bottom", fontsize=9, color="#333")
    if label_below:
        ax.text(x, y - radius - 0.22, label, ha="center", va="top",
                fontsize=9.5, color="#222")
    if status_badge:
        ax.text(x, y - radius - 0.65, status_badge, ha="center", va="top",
                fontsize=8, color="#1a7f37",
                bbox=dict(boxstyle="round,pad=0.25", facecolor="#dafbe1",
                          edgecolor="#1a7f37", linewidth=0.8))


def draw_arrow(ax, x1, y1, x2, y2, color, label=None, style="solid",
               radius=0.32):
    dx, dy = x2 - x1, y2 - y1
    dist = math.hypot(dx, dy)
    ux, uy = dx / dist, dy / dist
    sx, sy = x1 + ux * radius, y1 + uy * radius
    ex, ey = x2 - ux * radius, y2 - uy * radius
    ls = "--" if style == "dashed" else "-"
    arr = FancyArrowPatch(
        (sx, sy), (ex, ey),
        arrowstyle="->", mutation_scale=18,
        color=color, linewidth=2.2, linestyle=ls, zorder=2
    )
    ax.add_patch(arr)
    if label:
        mx, my = (sx + ex) / 2, (sy + ey) / 2
        ax.text(mx, my + 0.18, label, ha="center", va="bottom",
                fontsize=9.5, color=color, fontweight="bold")


# ---------------------------------------------------------------------------
# 2. Critical Path Analysis v6
# ---------------------------------------------------------------------------
fig, ax = plt.subplots(figsize=(20, 10), dpi=140)
ax.set_xlim(0, 20)
ax.set_ylim(0, 10)
ax.set_aspect("equal")
ax.axis("off")

# Title + UPDATED pill
ax.text(0.4, 9.55, "Critical Path Analysis  v6  (Apr 19, 2026)",
        fontsize=16, fontweight="bold", color="#111")
updated_pill = FancyBboxPatch((10.8, 9.35), 1.9, 0.45,
                              boxstyle="round,pad=0.02,rounding_size=0.1",
                              facecolor="#dafbe1", edgecolor="#1a7f37",
                              linewidth=1)
ax.add_patch(updated_pill)
ax.text(11.75, 9.58, "UPDATED", ha="center", va="center",
        fontsize=9.5, color="#1a7f37", fontweight="bold")

ax.text(19.6, 9.55,
        "FLOAT = 5 DAYS (NON-CRITICAL)  |  TOTAL CRITICAL PATH = 35 DAYS",
        ha="right", va="center", fontsize=9, color="#555")

# Legend box
legend_box = FancyBboxPatch((0.4, 7.05), 8.5, 1.95,
                            boxstyle="round,pad=0.05,rounding_size=0.12",
                            facecolor="#f6f8fb", edgecolor="#d0d7de",
                            linewidth=1)
ax.add_patch(legend_box)
ax.text(0.6, 8.78, "Start (Node 1)  \u2192  End (Node 12)",
        fontsize=10.5, fontweight="bold", color="#0d3b66")
ax.text(0.6, 8.50, "Path 1: Backend + NLP pipeline (main critical path)",
        fontsize=9.5, color="#222")
ax.text(0.6, 8.22, "Path 2: Frontend tasks (upper branch, non-critical)",
        fontsize=9.5, color="#2563eb")
ax.text(0.6, 7.94,
        "Path 3: Integration + Testing (E2E)  |  "
        "Path 4: Cloud Deploy (completed)",
        fontsize=9.5, color="#222")
ax.text(0.6, 7.66,
        "Path 5: QA / Demo  |  Path 6: Enhancement + Crash Testing",
        fontsize=9.5, color="#222")
ax.text(0.6, 7.38,
        "Path 7: Transactional Security Hardening (NEW, Apr 19)",
        fontsize=9.5, color="#222", fontweight="bold")

# Path pills
nc_blue = "#2563eb"
red = "#dc2626"
red_dark = "#991b1b"

p1_pill = FancyBboxPatch((0.7, 6.1), 4.1, 0.55,
                         boxstyle="round,pad=0.02,rounding_size=0.18",
                         facecolor="#eaf2ff", edgecolor=nc_blue, linewidth=1.6)
ax.add_patch(p1_pill)
ax.text(2.75, 6.38, "Path 1 (Non-Critical, Float = 5 days)",
        ha="center", va="center", fontsize=10, color=nc_blue,
        fontweight="bold")

p2_pill = FancyBboxPatch((5.1, 6.1), 3.8, 0.55,
                         boxstyle="round,pad=0.02,rounding_size=0.18",
                         facecolor="#fff5f5", edgecolor=red, linewidth=1.6)
ax.add_patch(p2_pill)
ax.text(7.0, 6.38, "Path 2 (Critical Path, Float = 0)",
        ha="center", va="center", fontsize=10, color=red,
        fontweight="bold")

# Node positions
main_y = 2.2
upper_y = 4.6

node_x = {
    1:  1.0,
    2:  2.8,
    4:  5.5,
    5:  7.8,
    7: 10.3,
    8: 12.4,
    9: 14.3,
    10: 16.2,
    11: 18.0,
    12: 19.2,
}
node_x[3] = 4.2   # Frontend (non-critical, upper)
node_x[6] = 8.9   # Cloud Deploy (deferred, upper)

# Main critical path nodes
crit_sequence = [1, 2, 4, 5, 7, 8, 9, 10, 11, 12]
crit_labels = {
    1:  "Start",
    2:  "Env & Consol.",
    4:  "Backend Val.",
    5:  "Data & RAG",
    7:  "QA & Demo",
    8:  "Freeze",
    9:  "Enhancement",
    10: "Adv. NLP",
    11: "Crash Test",
    12: "End",
}
for n in crit_sequence:
    draw_node(
        ax, node_x[n], main_y, crit_labels[n], str(n),
        color_face=red, color_edge=red_dark, text_color="white",
        status_badge="DONE",
    )

# Main critical path arrows with phase labels
crit_edges = [
    (1,  2,  "2d",  "Phase 1"),
    (2,  4,  "5d",  "Phase 2a: Backend"),
    (4,  5,  "5d",  "Phase 3: Data & RAG"),
    (5,  7,  "5d",  "Phase 5: QA & Demo"),
    (7,  8,  "5d",  "Phase 6: Freeze"),
    (8,  9,  "4d",  "Phase 7: Iteration"),
    (9,  10, "7d",  "Phase 8: Advanced NLP"),
    (10, 11, "2d",  "Phase 9: Testing"),
    (11, 12, "",    ""),
]
for a, b, dur, phase in crit_edges:
    draw_arrow(ax, node_x[a], main_y, node_x[b], main_y, red,
               label=dur if dur else None)
    if phase:
        mx = (node_x[a] + node_x[b]) / 2
        ax.text(mx, main_y + 0.60, phase, ha="center", va="bottom",
                fontsize=8, style="italic",
                color="#8a8a8a" if a == 1 else red_dark)

# Upper non-critical: Frontend (node 3)
draw_node(ax, node_x[3], upper_y, "Frontend", "3",
          color_face="white", color_edge=nc_blue, text_color=nc_blue,
          status_badge="DONE")
draw_arrow(ax, node_x[2], main_y, node_x[3], upper_y, nc_blue, label="5d")
ax.text((node_x[2] + node_x[3]) / 2 - 0.35,
        (main_y + upper_y) / 2 - 0.1,
        "Phase 2b: Frontend", ha="right", va="center",
        fontsize=8, style="italic", color=nc_blue)
draw_arrow(ax, node_x[3], upper_y, node_x[5], main_y, nc_blue,
           style="dashed")

# Upper: Cloud Deploy (node 6) — now DONE
deploy_green = "#1a7f37"
draw_node(ax, node_x[6], upper_y, "Cloud Deploy", "6",
          color_face="white", color_edge=deploy_green, text_color=deploy_green,
          status_badge="DONE")
draw_arrow(ax, node_x[5], main_y, node_x[6], upper_y, deploy_green, label="4d")
ax.text((node_x[5] + node_x[6]) / 2 + 0.35,
        (main_y + upper_y) / 2 - 0.35,
        "Phase 4: Cloud Deploy", ha="left", va="center",
        fontsize=8, style="italic", color=deploy_green)
draw_arrow(ax, node_x[6], upper_y, node_x[7], main_y, deploy_green,
           style="dashed")

# Updated sub-captions reflecting Apr 19 transactional security work
ax.text(node_x[8], main_y - 1.15,
        "Code Freeze\nCheckout Rate Limiting\nOrder Freezing",
        ha="center", va="top", fontsize=8.5, color="#555")

ax.text(node_x[9], main_y - 1.15,
        "Coreference  \u00b7  Preferences\n"
        "Memory  \u00b7  Re-ranking\n"
        "Txn Security  \u00b7  Idempotency",
        ha="center", va="top", fontsize=8.5, color="#555")

ax.text(node_x[10], main_y - 1.15,
        "Intent Classification\n"
        "Auth  \u00b7  Buttons  \u00b7  Logo\n"
        "Webhook Hardening",
        ha="center", va="top", fontsize=8.5, color="#555")

ax.text(node_x[11], main_y - 1.15,
        "Crash Tests\nDemo Prep",
        ha="center", va="top", fontsize=8.5, color="#555")

# Bottom legend strip
ax.add_patch(Circle((0.7, 0.5), 0.18, facecolor=red, edgecolor=red_dark,
                    linewidth=1.5))
ax.text(1.05, 0.5, "Critical Path", va="center", fontsize=9.5, color="#222")
ax.add_patch(Circle((3.2, 0.5), 0.18, facecolor="white", edgecolor=nc_blue,
                    linewidth=1.8))
ax.text(3.55, 0.5, "Non-Critical Path", va="center", fontsize=9.5,
        color="#222")
ax.text(6.0, 0.5, "DONE", va="center", fontsize=8.5, color="#1a7f37",
        bbox=dict(boxstyle="round,pad=0.25", facecolor="#dafbe1",
                  edgecolor="#1a7f37", linewidth=0.8))
ax.text(6.75, 0.5, "= Completed", va="center", fontsize=9.5, color="#222")
ax.add_patch(Circle((8.8, 0.5), 0.18, facecolor="white", edgecolor="#1a7f37",
                    linewidth=1.8))
ax.text(9.15, 0.5, "Completed (non-critical)", va="center", fontsize=9.5,
        color="#222")

plt.tight_layout()
crit_path = f"{DOCS}/Critical_Path_Analysis_v6.png"
plt.savefig(crit_path, dpi=160, bbox_inches="tight", facecolor="white")
plt.close(fig)
print(f"wrote {crit_path}")
