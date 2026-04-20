"""
Update Gantt xlsx, generate PERT chart PNG, generate Critical Path PNG
for the current state of the project as of 2026-04-18.
"""
import shutil
import datetime as dt
from copy import copy

import openpyxl
from openpyxl.styles import PatternFill, Font

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Circle, FancyArrowPatch, FancyBboxPatch
from matplotlib.lines import Line2D

ROOT = "/Users/billyerdely/Documents/Penn_State/Commerce-Agent-LLM"
DOCS = f"{ROOT}/docs"

# ---------------------------------------------------------------------------
# 1. Gantt: clone xlsx, mark Final Optimization & Documentation tasks Done
# ---------------------------------------------------------------------------
SRC = "/Users/billyerdely/Downloads/Gantt_Chart_v4 (1).xlsx"
DST = f"{DOCS}/Gantt_Chart_v5.xlsx"
shutil.copy(SRC, DST)

wb = openpyxl.load_workbook(DST)
ws = wb["Gantt Chart"]

# Update title
ws.cell(row=1, column=1).value = (
    "AI-Powered E-Commerce Shopping Assistant (Gantt Chart) — Updated v5"
)

done_fill = PatternFill(
    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
)

def set_done(row, completed_on, comment=None):
    ws.cell(row=row, column=5).value = "Done"
    ws.cell(row=row, column=7).value = completed_on
    ws.cell(row=row, column=5).fill = done_fill
    if comment is not None:
        ws.cell(row=row, column=8).value = comment

APR_10 = dt.datetime(2026, 4, 10)
APR_12 = dt.datetime(2026, 4, 12)
APR_15 = dt.datetime(2026, 4, 15)
APR_17 = dt.datetime(2026, 4, 17)
APR_18 = dt.datetime(2026, 4, 18)
APR_19 = dt.datetime(2026, 4, 19)

# Row 80: Phase summary
set_done(
    80,
    APR_18,
    "All optimization, security, docs, deploy, and regression work completed",
)
set_done(
    81,
    APR_10,
    "Embedding disk cache, lazy WS reconnect, markdown-to-jsx memoized; latency benchmarks logged",
)
set_done(
    82,
    APR_18,
    "JWT HS256 pinning + exp/sub/jti/iat, token denylist, /auth/guest rate limit, CSP headers, audit log, schema validators (test_phase1_security.py, 15 checks passing)",
)
set_done(
    83,
    APR_15,
    "MONITORING.md architecture flow + file map; README architecture section; crash_test_results.md",
)
set_done(
    84,
    APR_15,
    "START_HERE.md quickstart + example queries; demo walkthrough checklist",
)
set_done(
    85,
    APR_17,
    "v2.0 tagged; frontend and backend deployed with hardened auth and CORS locked to FRONTEND_URL",
)
set_done(
    86,
    APR_18,
    "Full regression run on prod: auth, cart merge, WS resiliency, category routing, Stripe sandbox checkout all passing",
)

# Final Presentation phase (rows 88 - 93): still in progress, today is Apr 18
# Mark the parent and deck prep as In Progress to reflect honest state
ws.cell(row=88, column=5).value = "In Progress"
ws.cell(row=88, column=8).value = (
    "Deck build underway; delivery scheduled 2026-04-28, submission 2026-05-01"
)
ws.cell(row=89, column=5).value = "In Progress"
ws.cell(row=89, column=8).value = (
    "Deck drafted using Results_Conclusions_Future_Work.docx as backbone"
)

wb.save(DST)
print(f"wrote {DST}")


# ---------------------------------------------------------------------------
# Shared drawing helpers
# ---------------------------------------------------------------------------
def draw_node(ax, x, y, label, number, color_face, color_edge, text_color,
              radius=0.32, label_below=True, label_above_text=None,
              status_badge=None):
    c = Circle((x, y), radius, facecolor=color_face, edgecolor=color_edge,
               linewidth=2.5, zorder=3)
    ax.add_patch(c)
    ax.text(x, y, str(number), ha="center", va="center",
            fontsize=13, fontweight="bold", color=text_color, zorder=4)
    if label_above_text:
        ax.text(x, y + radius + 0.28, label_above_text, ha="center",
                va="bottom", fontsize=9, color="#333")
    if label_below:
        ax.text(x, y - radius - 0.22, label, ha="center", va="top",
                fontsize=9.5, color="#222")
    if status_badge:
        ax.text(x, y - radius - 0.65, status_badge, ha="center", va="top",
                fontsize=8, color="#1a7f37",
                bbox=dict(boxstyle="round,pad=0.25", facecolor="#dafbe1",
                          edgecolor="#1a7f37", linewidth=0.8))

def draw_arrow(ax, x1, y1, x2, y2, color, label=None, style="solid",
               radius=0.32):
    # shorten so arrow doesn't overlap node
    import math
    dx, dy = x2 - x1, y2 - y1
    dist = math.hypot(dx, dy)
    ux, uy = dx / dist, dy / dist
    sx, sy = x1 + ux * radius, y1 + uy * radius
    ex, ey = x2 - ux * radius, y2 - uy * radius
    ls = "--" if style == "dashed" else "-"
    arr = FancyArrowPatch(
        (sx, sy), (ex, ey),
        arrowstyle="->", mutation_scale=18,
        color=color, linewidth=2.2, linestyle=ls, zorder=2
    )
    ax.add_patch(arr)
    if label:
        mx, my = (sx + ex) / 2, (sy + ey) / 2
        ax.text(mx, my + 0.18, label, ha="center", va="bottom",
                fontsize=9.5, color=color, fontweight="bold")


# ---------------------------------------------------------------------------
# 2. PERT Chart
# ---------------------------------------------------------------------------
fig, ax = plt.subplots(figsize=(18, 9), dpi=140)
ax.set_xlim(0, 18)
ax.set_ylim(0, 9)
ax.set_aspect("equal")
ax.axis("off")

# Title
ax.text(9, 8.5, "AI Shopping Assistant  —  PERT Chart (Final State)",
        ha="center", va="center", fontsize=16, fontweight="bold",
        color="#111")

# Legend box (upper left)
legend = FancyBboxPatch((0.3, 6.9), 5.3, 1.4,
                        boxstyle="round,pad=0.05,rounding_size=0.12",
                        facecolor="#f6f8fb", edgecolor="#d0d7de",
                        linewidth=1)
ax.add_patch(legend)
ax.text(0.5, 8.05, "Start (Node 1)  →  End (Node 13)", fontsize=10.5,
        fontweight="bold", color="#0d3b66")
ax.text(0.5, 7.75, "Path 1: NLP pipeline enhancements (main path)",
        fontsize=9.5, color="#222")
ax.text(0.5, 7.5, "Path 2: UI / Auth enhancements (upper branch)",
        fontsize=9.5, color="#3a86ff")
ax.text(0.5, 7.25, "Path 3: Final optimization, security, and deploy",
        fontsize=9.5, color="#222")

# Main (bottom) path coords: 1 .. 13
main_y = 3.2
main_xs = [1.0, 2.5, 4.0, 5.5, 7.0, 8.5, 10.0, 11.5, 13.0, 14.5, 16.0, 17.0]
# That's 12 x positions for 12 nodes; End is at x=17.0 node 13
main_labels = [
    ("Start", "1"),
    ("Coreference\nResolution", "2"),
    ("Preference\nExtraction", "3"),
    ("Memory\nSummarization", "4"),
    ("Search\nRe-ranking", "5"),
    ("API Priority\nFix", "6"),
    ("Merge &\nValidate", "7"),
    ("Security\nAudit", "8"),
    ("Documentation", "9"),
    ("Production\nDeploy", "10"),
    ("Final\nRegression", "11"),
    ("End", "12"),
]

dark = "#14213d"
for (lbl, num), x in zip(main_labels, main_xs):
    is_end = (num in ("1", "12"))
    draw_node(
        ax, x, main_y, lbl, num,
        color_face=dark if is_end else "white",
        color_edge=dark,
        text_color="white" if is_end else "#14213d",
    )

# Main path arrows with durations
main_durations = ["2d", "2d", "1d", "1d", "1d", "1d", "5d", "5d", "3d", "3d",
                  "1d"]
for i in range(len(main_xs) - 1):
    draw_arrow(ax, main_xs[i], main_y, main_xs[i + 1], main_y, "#14213d",
               label=main_durations[i])

# Upper (UI/Auth) branch: B1, B2, B3
blue = "#3a86ff"
b_y = 5.6
b_xs = [4.0, 6.2, 8.4]
for (lbl, tag), x in zip(
    [("Login / Auth\nPage", "B1"),
     ("Button\nFunctionality", "B2"),
     ("Logo &\nBranding", "B3")], b_xs):
    draw_node(ax, x, b_y, lbl, tag, "white", blue, blue)

# 1 -> B1
draw_arrow(ax, main_xs[0], main_y, b_xs[0], b_y, blue, label="2d")
draw_arrow(ax, b_xs[0], b_y, b_xs[1], b_y, blue, label="1d")
draw_arrow(ax, b_xs[1], b_y, b_xs[2], b_y, blue, label="1d")
# B3 dashed join back to node 7 (Merge & Validate)
draw_arrow(ax, b_xs[2], b_y, main_xs[6], main_y, blue, label="join",
           style="dashed")

# Sub-captions under main path
ax.text((main_xs[1] + main_xs[5]) / 2, 2.15,
        "CoreferenceResolver  ·  extract_and_update_preferences()  ·  "
        "_summarize_messages()\nPreference boosting on RRF  ·  "
        "SerpAPI → RapidAPI → eBay → local priority",
        ha="center", va="top", fontsize=8.5, color="#555")

ax.text((main_xs[7] + main_xs[10]) / 2, 2.15,
        "JWT HS256 pinning  ·  CSP + rate limits  ·  MONITORING.md  ·  "
        "v2.0 prod deploy\n15 security checks  ·  Full regression on prod",
        ha="center", va="top", fontsize=8.5, color="#555")

# Sub-caption for upper branch
ax.text((b_xs[0] + b_xs[2]) / 2, 6.35,
        "OAuth (Azure MSAL + Google)\n"
        "All button handlers wired  ·  New gradient logo",
        ha="center", va="bottom", fontsize=8.5, color="#555")

plt.tight_layout()
pert_path = f"{DOCS}/PERT_Chart_v5.png"
plt.savefig(pert_path, dpi=160, bbox_inches="tight", facecolor="white")
plt.close(fig)
print(f"wrote {pert_path}")


# ---------------------------------------------------------------------------
# 3. Critical Path Analysis (phase-level view, distinct from PERT)
# ---------------------------------------------------------------------------
fig, ax = plt.subplots(figsize=(20, 10), dpi=140)
ax.set_xlim(0, 20)
ax.set_ylim(0, 10)
ax.set_aspect("equal")
ax.axis("off")

# Title + UPDATED pill
ax.text(0.4, 9.55, "Critical Path Analysis  —  AI Shopping Assistant Project",
        fontsize=16, fontweight="bold", color="#111")
updated_pill = FancyBboxPatch((12.35, 9.35), 1.9, 0.45,
                              boxstyle="round,pad=0.02,rounding_size=0.1",
                              facecolor="#dafbe1", edgecolor="#1a7f37",
                              linewidth=1)
ax.add_patch(updated_pill)
ax.text(13.3, 9.58, "UPDATED", ha="center", va="center",
        fontsize=9.5, color="#1a7f37", fontweight="bold")

ax.text(19.6, 9.55,
        "FLOAT = 5 DAYS (NON-CRITICAL)  ·  TOTAL CRITICAL PATH = 35 DAYS",
        ha="right", va="center", fontsize=9, color="#555")

# Legend box (upper-left) listing all 6 paths
legend_box = FancyBboxPatch((0.4, 7.25), 8.0, 1.75,
                            boxstyle="round,pad=0.05,rounding_size=0.12",
                            facecolor="#f6f8fb", edgecolor="#d0d7de",
                            linewidth=1)
ax.add_patch(legend_box)
ax.text(0.6, 8.78, "Start (Node 1)  →  End (Node 12)",
        fontsize=10.5, fontweight="bold", color="#0d3b66")
ax.text(0.6, 8.50, "Path 1: Backend + NLP pipeline (main critical path)",
        fontsize=9.5, color="#222")
ax.text(0.6, 8.22, "Path 2: Frontend tasks (upper branch, non-critical)",
        fontsize=9.5, color="#2563eb")
ax.text(0.6, 7.94,
        "Path 3: Integration + Testing (E2E)  ·  "
        "Path 4: Cloud Deploy (deferred)",
        fontsize=9.5, color="#222")
ax.text(0.6, 7.66,
        "Path 5: QA / Demo  ·  Path 6: Enhancement + Crash Testing (NEW)",
        fontsize=9.5, color="#222", fontweight="bold")

# Two path pills (below legend)
nc_blue = "#2563eb"
red = "#dc2626"
red_dark = "#991b1b"

p1_pill = FancyBboxPatch((0.7, 6.3), 4.1, 0.55,
                         boxstyle="round,pad=0.02,rounding_size=0.18",
                         facecolor="#eaf2ff", edgecolor=nc_blue, linewidth=1.6)
ax.add_patch(p1_pill)
ax.text(2.75, 6.58, "Path 1 (Non-Critical, Float = 5 days)",
        ha="center", va="center", fontsize=10, color=nc_blue,
        fontweight="bold")

p2_pill = FancyBboxPatch((5.1, 6.3), 3.8, 0.55,
                         boxstyle="round,pad=0.02,rounding_size=0.18",
                         facecolor="#fff5f5", edgecolor=red, linewidth=1.6)
ax.add_patch(p2_pill)
ax.text(7.0, 6.58, "Path 2 (Critical Path, Float = 0)",
        ha="center", va="center", fontsize=10, color=red,
        fontweight="bold")

# Main critical path (lower, red) and branches
main_y = 2.2

# Node x positions 1..12 along main path. Nodes 3 (Frontend) and 6 (Cloud
# Deploy) live on the upper row, so we leave their slots empty on the main row.
node_x = {
    1:  1.0,
    2:  2.8,
    4:  5.5,
    5:  7.8,
    7: 10.3,
    8: 12.4,
    9: 14.3,
    10: 16.2,
    11: 18.0,
    12: 19.2,
}
upper_y = 4.6
node_x[3] = 4.2      # Frontend (non-critical, upper)
node_x[6] = 8.9      # Cloud Deploy (deferred, upper)

# Main critical path nodes (red, DONE)
crit_sequence = [1, 2, 4, 5, 7, 8, 9, 10, 11, 12]
crit_labels = {
    1:  "Start",
    2:  "Env & Consol.",
    4:  "Backend Val.",
    5:  "Data & RAG",
    7:  "QA & Demo",
    8:  "Freeze",
    9:  "Enhancement",
    10: "Adv. NLP",
    11: "Crash Test",
    12: "End",
}
for n in crit_sequence:
    draw_node(
        ax, node_x[n], main_y, crit_labels[n], str(n),
        color_face=red, color_edge=red_dark, text_color="white",
        status_badge="DONE",
    )

# Main critical path arrows with phase labels
crit_edges = [
    (1,  2,  "2d",  "Phase 1"),
    (2,  4,  "5d",  "Phase 2a: Backend"),
    (4,  5,  "5d",  "Phase 3: Data & RAG"),
    (5,  7,  "5d",  "Phase 5: QA & Demo"),
    (7,  8,  "5d",  "Phase 6: Freeze"),
    (8,  9,  "4d",  "Phase 7: Iteration"),
    (9,  10, "7d",  "Phase 8: Advanced NLP"),
    (10, 11, "2d",  "Phase 9: Testing"),
    (11, 12, "",    ""),
]
for a, b, dur, phase in crit_edges:
    draw_arrow(ax, node_x[a], main_y, node_x[b], main_y, red,
               label=dur if dur else None)
    if phase:
        mx = (node_x[a] + node_x[b]) / 2
        ax.text(mx, main_y + 0.60, phase, ha="center", va="bottom",
                fontsize=8, style="italic",
                color="#8a8a8a" if a == 1 else red_dark)

# Upper non-critical: Frontend (node 3)
draw_node(ax, node_x[3], upper_y, "Frontend", "3",
          color_face="white", color_edge=nc_blue, text_color=nc_blue,
          status_badge="DONE")
draw_arrow(ax, node_x[2], main_y, node_x[3], upper_y, nc_blue, label="5d")
ax.text((node_x[2] + node_x[3]) / 2 - 0.35,
        (main_y + upper_y) / 2 - 0.1,
        "Phase 2b: Frontend", ha="right", va="center",
        fontsize=8, style="italic", color=nc_blue)
# Dashed rejoin to node 5
draw_arrow(ax, node_x[3], upper_y, node_x[5], main_y, nc_blue,
           style="dashed")

# Upper deferred: Cloud Deploy (node 6)
gray = "#9aa0a6"
c6 = Circle((node_x[6], upper_y), 0.32, facecolor="white",
            edgecolor=gray, linewidth=2, linestyle="--", zorder=3)
ax.add_patch(c6)
ax.text(node_x[6], upper_y, "6", ha="center", va="center",
        fontsize=13, fontweight="bold", color=gray, zorder=4)
ax.text(node_x[6], upper_y + 0.55, "Cloud Deploy",
        ha="center", va="bottom", fontsize=9.5, color=gray)
defer_pill = FancyBboxPatch(
    (node_x[6] - 0.42, upper_y + 0.9), 0.85, 0.32,
    boxstyle="round,pad=0.02,rounding_size=0.1",
    facecolor="#fff4c2", edgecolor="#a47a00", linewidth=0.8)
ax.add_patch(defer_pill)
ax.text(node_x[6], upper_y + 1.06, "DEFER",
        ha="center", va="center", fontsize=8, color="#7a5a00",
        fontweight="bold")
# Dashed gray arrow 5 -> 6
draw_arrow(ax, node_x[5], main_y, node_x[6], upper_y, gray,
           label="4d", style="dashed")
ax.text((node_x[5] + node_x[6]) / 2 + 0.35,
        (main_y + upper_y) / 2 - 0.35,
        "Phase 4: Cloud Deploy", ha="left", va="center",
        fontsize=8, style="italic", color=gray)

# Captions under enhancement / adv NLP
ax.text(node_x[9], main_y - 1.15,
        "Coreference  ·  Preferences\nMemory  ·  Re-ranking",
        ha="center", va="top", fontsize=8.5, color="#555")
ax.text(node_x[10], main_y - 1.15,
        "Intent Classification\nAuth  ·  Buttons  ·  Logo",
        ha="center", va="top", fontsize=8.5, color="#555")

# Bottom legend strip
ax.add_patch(Circle((0.7, 0.5), 0.18, facecolor=red, edgecolor=red_dark,
                    linewidth=1.5))
ax.text(1.05, 0.5, "Critical Path", va="center", fontsize=9.5, color="#222")
ax.add_patch(Circle((3.2, 0.5), 0.18, facecolor="white", edgecolor=nc_blue,
                    linewidth=1.8))
ax.text(3.55, 0.5, "Non-Critical Path", va="center", fontsize=9.5,
        color="#222")
ax.text(6.0, 0.5, "DONE", va="center", fontsize=8.5, color="#1a7f37",
        bbox=dict(boxstyle="round,pad=0.25", facecolor="#dafbe1",
                  edgecolor="#1a7f37", linewidth=0.8))
ax.text(6.75, 0.5, "= Completed", va="center", fontsize=9.5, color="#222")
ax.text(8.8, 0.5, "DEFER", va="center", fontsize=8, color="#7a5a00",
        fontweight="bold",
        bbox=dict(boxstyle="round,pad=0.25", facecolor="#fff4c2",
                  edgecolor="#a47a00", linewidth=0.8))
ax.text(9.6, 0.5, "= Deferred", va="center", fontsize=9.5, color="#222")

plt.tight_layout()
crit_path = f"{DOCS}/Critical_Path_Analysis_v5.png"
plt.savefig(crit_path, dpi=160, bbox_inches="tight", facecolor="white")
plt.close(fig)
print(f"wrote {crit_path}")
